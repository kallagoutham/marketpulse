"""Stream stock market CSV rows to Kafka as JSON events."""

from __future__ import annotations

import argparse
import csv
import json
import os
import signal
import socket
import sys
import time
from pathlib import Path
from typing import Any, Iterable

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATASET = BASE_DIR / "dataset" / "indexProcessed.csv"

NUMERIC_FIELDS = {
    "Open",
    "High",
    "Low",
    "Close",
    "Adj Close",
    "Volume",
    "CloseUSD",
}

REQUIRED_FIELDS = {
    "Index",
    "Date",
    "Open",
    "High",
    "Low",
    "Close",
    "Adj Close",
    "Volume",
    "CloseUSD",
}

stop_requested = False


def request_stop(signum: int, frame: object) -> None:
    del signum, frame
    global stop_requested
    stop_requested = True


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Publish dataset/indexProcessed.csv rows to Kafka."
    )
    parser.add_argument(
        "--dataset",
        default=os.environ.get("STOCK_DATASET_PATH", str(DEFAULT_DATASET)),
        help="Path to the CSV dataset.",
    )
    parser.add_argument(
        "--bootstrap-servers",
        default=os.environ.get("KAFKA_BOOTSTRAP_SERVERS"),
        help="Kafka bootstrap servers. Defaults to KAFKA_BOOTSTRAP_SERVERS from .env.",
    )
    parser.add_argument(
        "--topic",
        default=os.environ.get("KAFKA_TOPIC"),
        help="Kafka topic to publish messages to. Defaults to KAFKA_TOPIC from .env.",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=float(os.environ.get("PRODUCER_DELAY_SECONDS", "0.2")),
        help="Seconds to wait between messages.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=int(os.environ.get("PRODUCER_LIMIT", "0")),
        help="Maximum number of rows to publish. Use 0 for all rows.",
    )
    parser.add_argument(
        "--loop",
        action="store_true",
        default=os.environ.get("PRODUCER_LOOP", "False").lower() == "true",
        help="Continuously replay the dataset until stopped.",
    )
    parser.add_argument(
        "--create-topic",
        action=argparse.BooleanOptionalAction,
        default=os.environ.get("KAFKA_CREATE_TOPIC", "True").lower() == "true",
        help="Create the Kafka topic if it does not exist. Enabled by default.",
    )
    parser.add_argument(
        "--partitions",
        type=int,
        default=int(os.environ.get("KAFKA_TOPIC_PARTITIONS", "1")),
        help="Partition count to use when creating the topic.",
    )
    parser.add_argument(
        "--replication-factor",
        type=int,
        default=int(os.environ.get("KAFKA_TOPIC_REPLICATION_FACTOR", "1")),
        help="Replication factor to use when creating the topic.",
    )
    return parser.parse_args()


def normalize_value(field: str, value: Any) -> Any:
    if value is None or value == "":
        return None
    if field in NUMERIC_FIELDS:
        return float(value)
    return value


def validate_fields(fields: Iterable[str], dataset_path: Path) -> None:
    present_fields = set(fields)
    missing_fields = sorted(REQUIRED_FIELDS - present_fields)
    if missing_fields:
        raise ValueError(
            f"{dataset_path} is missing required columns: {', '.join(missing_fields)}"
        )


def normalize_row(row: dict[str, Any], source: str) -> dict[str, Any]:
    return {
        "index": row["Index"],
        "date": row["Date"],
        "open": normalize_value("Open", row["Open"]),
        "high": normalize_value("High", row["High"]),
        "low": normalize_value("Low", row["Low"]),
        "close": normalize_value("Close", row["Close"]),
        "adj_close": normalize_value("Adj Close", row["Adj Close"]),
        "volume": normalize_value("Volume", row["Volume"]),
        "close_usd": normalize_value("CloseUSD", row["CloseUSD"]),
        "source": source,
        "event_type": "stock_market_tick",
        "published_at": int(time.time()),
    }


def iter_csv_events(dataset_path: Path):
    with dataset_path.open("r", encoding="utf-8", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        validate_fields(reader.fieldnames or [], dataset_path)
        for row in reader:
            yield normalize_row(row, source=str(dataset_path))


def iter_xlsx_events(dataset_path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency: openpyxl. Install dependencies with "
            "`pip install -r requirements.txt`."
        ) from exc

    workbook = load_workbook(dataset_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    validate_fields(headers, dataset_path)

    for values in rows:
        row = dict(zip(headers, values, strict=False))
        yield normalize_row(row, source=str(dataset_path))


def iter_events(dataset_path: Path):
    suffix = dataset_path.suffix.lower()
    if suffix == ".csv":
        yield from iter_csv_events(dataset_path)
    elif suffix in {".xlsx", ".xlsm"}:
        yield from iter_xlsx_events(dataset_path)
    else:
        raise ValueError("Dataset must be a .csv, .xlsx, or .xlsm file.")


def build_producer(bootstrap_servers: str):
    try:
        from kafka import KafkaProducer
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency: kafka-python. Install dependencies with "
            "`pip install -r requirements.txt`."
        ) from exc

    return KafkaProducer(
        bootstrap_servers=[server.strip() for server in bootstrap_servers.split(",")],
        key_serializer=lambda key: key.encode("utf-8"),
        value_serializer=lambda value: json.dumps(value).encode("utf-8"),
        acks="all",
        retries=3,
        linger_ms=10,
    )


def ensure_topic(
    bootstrap_servers: str,
    topic: str,
    partitions: int,
    replication_factor: int,
) -> None:
    try:
        from kafka.admin import KafkaAdminClient, NewTopic
        from kafka.errors import TopicAlreadyExistsError
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency: kafka-python. Install dependencies with "
            "`pip install -r requirements.txt`."
        ) from exc

    admin = KafkaAdminClient(
        bootstrap_servers=[server.strip() for server in bootstrap_servers.split(",")],
        client_id="marketpulse-topic-admin",
    )

    try:
        existing_topics = set(admin.list_topics())
        if topic in existing_topics:
            print(f"Kafka topic already exists: {topic}")
            return

        admin.create_topics(
            [
                NewTopic(
                    name=topic,
                    num_partitions=partitions,
                    replication_factor=replication_factor,
                )
            ]
        )
        print(f"Created Kafka topic: {topic}")
    except TopicAlreadyExistsError:
        print(f"Kafka topic already exists: {topic}")
    finally:
        admin.close()


def check_bootstrap_servers(
    bootstrap_servers: str, timeout_seconds: float = 5.0
) -> bool:
    reachable = False

    for server in [
        item.strip() for item in bootstrap_servers.split(",") if item.strip()
    ]:
        host, separator, port_text = server.rpartition(":")
        if not separator or not host or not port_text:
            print(f"Invalid Kafka bootstrap server value: {server}", file=sys.stderr)
            continue

        try:
            with socket.create_connection(
                (host, int(port_text)), timeout=timeout_seconds
            ):
                reachable = True
        except OSError as exc:
            print(f"Cannot reach Kafka broker {server}: {exc}", file=sys.stderr)

    return reachable


def publish_events(
    producer,
    topic: str,
    dataset_path: Path,
    delay_seconds: float,
    limit: int,
    progress_callback=None,
) -> int:
    sent_count = 0
    for event in iter_events(dataset_path):
        if stop_requested:
            break
        if limit and sent_count >= limit:
            break

        producer.send(topic, key=event["index"], value=event)
        sent_count += 1
        if sent_count == 1:
            print(f"Publishing first message to {topic}")

        if sent_count % 100 == 0:
            producer.flush()
            print(f"Published {sent_count} messages to {topic}")
            if progress_callback:
                progress_callback(sent_count)

        if delay_seconds > 0:
            time.sleep(delay_seconds)

    producer.flush()
    if progress_callback:
        progress_callback(sent_count)
    return sent_count


def main() -> int:
    if load_dotenv:
        load_dotenv(BASE_DIR / ".env")

    signal.signal(signal.SIGINT, request_stop)
    signal.signal(signal.SIGTERM, request_stop)

    args = parse_args()
    dataset_path = Path(args.dataset).expanduser().resolve()

    if not args.bootstrap_servers:
        print(
            "KAFKA_BOOTSTRAP_SERVERS is required in .env or --bootstrap-servers.",
            file=sys.stderr,
        )
        return 1

    if not args.topic:
        print("KAFKA_TOPIC is required in .env or --topic.", file=sys.stderr)
        return 1

    if not dataset_path.exists():
        print(f"Dataset not found: {dataset_path}", file=sys.stderr)
        return 1

    print(f"Connecting to Kafka broker: {args.bootstrap_servers}")
    print(f"Publishing dataset: {dataset_path}")
    print(f"Kafka topic: {args.topic}")

    if not check_bootstrap_servers(args.bootstrap_servers):
        print(
            "Kafka bootstrap check failed. Verify the EC2 security group, Docker port "
            "mapping, and Kafka advertised.listeners setting.",
            file=sys.stderr,
        )
        return 1

    if args.create_topic:
        ensure_topic(
            bootstrap_servers=args.bootstrap_servers,
            topic=args.topic,
            partitions=args.partitions,
            replication_factor=args.replication_factor,
        )

    producer = build_producer(args.bootstrap_servers)
    total_sent = 0

    try:
        while True:
            sent_count = publish_events(
                producer=producer,
                topic=args.topic,
                dataset_path=dataset_path,
                delay_seconds=args.delay,
                limit=args.limit,
            )
            total_sent += sent_count
            print(f"Finished pass. Published {sent_count} messages.")

            if not args.loop or stop_requested:
                break
    except Exception as exc:
        print(f"Kafka publish failed: {exc}", file=sys.stderr)
        return 1
    finally:
        producer.close()

    print(f"Producer stopped. Total messages published: {total_sent}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
